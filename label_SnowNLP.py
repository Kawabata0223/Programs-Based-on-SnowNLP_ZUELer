import openpyxl
from snownlp import SnowNLP
import pandas as pd


workbook = openpyxl.load_workbook('files/comments.xlsx')

sheet = workbook.active

max_row = sheet.max_row


for row in range(2, max_row + 1):
    cell_value = sheet.cell(row=row, column=1).value

    # 使用 SnowNLP 进行情感分析
    s = SnowNLP(cell_value)
    sentiment_score = s.sentiments

    # 将情感分数存储到第二列
    sheet.cell(row=row, column=2, value=sentiment_score)


workbook.save('labels/comments_score.xlsx')

workbook.close()

#--------------------------------------------------------

input_excel_path = 'labels/comments_score.xlsx'
df = pd.read_excel(input_excel_path)


filtered_df = df[(df.iloc[:, 1] >= 0) & (df.iloc[:, 1] < 0.2)]
output_excel_path = 'labels/label1.xlsx'
filtered_df.to_excel(output_excel_path, index=False)


filtered_df = df[(df.iloc[:, 1] >= 0.2) & (df.iloc[:, 1] < 0.4)]
output_excel_path = 'labels/label2.xlsx'
filtered_df.to_excel(output_excel_path, index=False)



filtered_df = df[(df.iloc[:, 1] >= 0.4) & (df.iloc[:, 1] < 0.6)]
output_excel_path = 'labels/label3.xlsx'
filtered_df.to_excel(output_excel_path, index=False)


filtered_df = df[(df.iloc[:, 1] >= 0.6) & (df.iloc[:, 1] < 0.8)]
output_excel_path = 'labels/label4.xlsx'
filtered_df.to_excel(output_excel_path, index=False)


filtered_df = df[(df.iloc[:, 1] >= 0.8) & (df.iloc[:, 1] <= 1)]
output_excel_path = 'labels/label5.xlsx'
filtered_df.to_excel(output_excel_path, index=False)


print("情感分析完成")


